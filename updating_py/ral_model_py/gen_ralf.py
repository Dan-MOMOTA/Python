import os
import openpyxl

dst_filename = 'mgc_uart_reg.ralf'
workbook = openpyxl.load_workbook('REG.xlsx')  # 返回一个workbook数据类型的值
file_obj = open(dst_filename,'w') # 打开目标文件
print(workbook.sheetnames)  # 打印Excel表中的所有表
sheet = workbook['Sheet1']  # 获取某一个表
print(sheet.dimensions)     # 获取表格的尺寸大小
row_num = sheet.max_row     # 行
line_num = sheet.max_column # 列

reg_column          = 'A'   # 各部分的所在列
address_column      = 'B'
reg_access_column   = 'C'
field_column        = 'D'
field_access_column = 'E'
reset_value_column  = 'F'
bit_end_column      = 'G'
bit_start_column    = 'H'
function_column     = 'I'
# 这里只有一个block所在在循环外面写
file_obj.write("block uart_reg_block { \n")
file_obj.write("    bytes " + str(11)+';\n') # byte 
for i in range(2,row_num+1): #  i < row_num 所以要加一 第一行是描述性的
    j = i
    cell_data = sheet[reg_column + str(i)]
    if(cell_data.value!=None): # 处理寄存器域
        start_row = cell_data.row
        # 寄存器的相关属性
        reg_name = cell_data.value
        reg_addr = sheet[address_column + str(i)].value
        reg_access = sheet[reg_access_column + str(i)].value
        file_obj.write("    register " + reg_name+ " @" +str(reg_addr)+" {\n")
        # file_obj.write("bytes " + str(1)) # byte 没有则自动计算
        print("reg name is",reg_name,"reg addr is ",reg_addr,"reg access is ",reg_access)
        while(1):
            j += 1
            if(j < row_num):
                if(sheet[reg_column + str(j)].value!=None):
                    end_row = j 
                    reg_filed = end_row - start_row 
                    print(cell_data.value,"reg  filed is" ,reg_filed )
                    break
            else:
                end_row = j + 1
                reg_filed = end_row - start_row
                print(cell_data.value,"reg  filed is" ,reg_filed )
                break

        for field_num in range(0,reg_filed) :
            fiedl_name = sheet[field_column+str(i+field_num)].value                 # 域的名字
            print("fiedl_name is ",fiedl_name)
            field_access = sheet[field_access_column+str(i+field_num)].value        # 域的access
            print("field access is ",field_access)
            field_reset_num_str = sheet[reset_value_column + str(i+field_num)].value# 复位值
            field_reset_num =  str(field_reset_num_str)[2:]  
            print("field_reset_nume ",field_reset_num)
            field_bit_start = sheet[bit_start_column+str(i+field_num)].value        # 起始位数
            field_bit_end = sheet[bit_end_column+str(i+field_num)].value            # 最终位数
            field_bits = field_bit_end - field_bit_start + 1                        # 共计多少bit
            print("bit start " , field_bit_start , "end " ,field_bit_end ,"bits " ,field_bits)
            file_obj.write("        field " + str(fiedl_name)+" {\n")
            file_obj.write("            bits " + str(field_bits)+";\n")
            file_obj.write("            reset " + str(field_reset_num)+";\n")
            file_obj.write("            access " + str(field_access).lower()+";\n") # ralgen 区分大小写
            file_obj.write("        }\n")
        file_obj.write("    }\n")   
    print("out for ",i)
file_obj.write("}")
file_obj.close()
print("clear")
SystemExit
